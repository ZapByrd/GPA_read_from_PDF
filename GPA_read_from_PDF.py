# 适用于西建大教务处系统上下载下来的成绩单.pdf，用以简易GPA计算  ————刘泽鹏，2023/8/24完成
# 读取pdf文件
import pdfplumber
import pandas as pd
import openpyxl
import time
T1 = time.perf_counter()  # 计算程序运行时间
# 读取pdf数据
pdf = pdfplumber.open('学生成绩.pdf')
pages = pdf.pages
print(f'共读取{len(pages)}页表格数据')
# 找到每页数据所属的学期
terms = []  # 列表，代表学期
i = 0
for page in pages:
    content = page.extract_text()
    if page.chars[0]['size'] > 14:
        st = ''  # 空字符串
        for j in range(0, 11):  # 找学期号码，为前11位占位符
            st = st + page.chars[j]['text']
        terms.append(st)
    else:
        terms.append(terms[i-1])
    i = i + 1
# 将表格数据提取至excel并计算
gpa, total, total_cdt, name_min, name_max = [], [], [], [], []
min_g, max_g, q = 100, 0, 0
is_con = 1  # 是否继续下步循环
for index, term in enumerate(terms):
    if is_con == 1:
        t = term
        if index < len(terms)-1 and terms[index+1] == term:  # 多页读取
            is_con = 0
            tables = []
            for page in pages[index:index+2]:
                table = page.extract_table()
                tables.extend(table)
        else:               # 单页读取
            page = pages[index]
            tables = page.extract_table()
            is_con = 1
        data = pd.DataFrame(tables[:],)  # 适用于一页一个表格
        data = data.replace({"": None})
        data = data.dropna()
        data.to_excel(f'{t}成绩单.xlsx', index=False)
# 计算
        wb = openpyxl.load_workbook(f'{t}成绩单.xlsx')
        sheet = wb['Sheet1']
        AllData = list(sheet.values)  # 表中的所有数据，以行为单位
        sum_a, sum_b = 0, 0
        for p in range(3, len(AllData)+1):
            cdt = float(sheet.cell(row=p, column=4).value)  # 学分
            gp = float(sheet.cell(row=p, column=5).value)   # 绩点
            if gp <= min_g:     # 最低分
                if gp < min_g:
                    min_g = gp
                    del name_min[:]
                name_min.append(sheet.cell(row=p, column=1).value)
            if gp >= max_g:     # 最高分
                if gp > max_g:
                    max_g = gp
                    del name_max[:]
                name_max.append(sheet.cell(row=p, column=1).value)
            sum_a = sum_a + cdt * gp
            sum_b = sum_b + cdt
        total.append(sum_a)
        total_cdt.append(sum_b)
        gpa.append(total[q] / total_cdt[q])
        print(f'已生成{t}成绩单.xlsx,该学期总学分为{total_cdt[q]},平均绩点为{round(gpa[q],4)}')
        q = q + 1
    else:
        is_con = 1
# 输出总的结果
print(f'所有学期总的学分为{sum(total_cdt)},平均绩点为{round(sum(total) / sum(total_cdt), 4)},'
      f'最高分为{name_max}——{max_g*10+50},最低分为{name_min}——{min_g*10+50}')
pdf.close()
# 输出运行时间
T2 = time.perf_counter()
print(f'程序用时为{T2-T1}秒')
